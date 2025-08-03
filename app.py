""import gradio as gr
import pandas as pd
import joblib
import numpy as np
import shap
import matplotlib.pyplot as plt
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.linear_model import LinearRegression
from fpdf import FPDF
import os
from datetime import datetime
import openpyxl

# Load model and scaler
model = joblib.load("model.pkl")
scaler = joblib.load("scaler.pkl")

# Email of developer
DEVELOPER_EMAIL = "arslanhafeezkhan.16@gmail.com"
LOG_FILE = "user_logs.xlsx"

# Feature columns
features = ["Cement", "Sand", "Coarse Aggregate", "Water", "Superplasticizer", "Fly Ash", "Slag", "Age"]

# ML Models for ensemble
rf = RandomForestRegressor(n_estimators=100, random_state=42)
gb = GradientBoostingRegressor(n_estimators=100, random_state=42)
lr = LinearRegression()

# Create SHAP explainer
def get_shap_values(input_df):
    rf.fit(input_df, model.predict(input_df))
    explainer = shap.Explainer(rf)
    shap_values = explainer(input_df)
    return shap_values

# Save user logs
def log_prediction(email, input_data, prediction):
    data = input_data.copy()
    data['Email'] = email
    data['Prediction'] = prediction
    data['Timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    df = pd.DataFrame([data])
    if not os.path.exists(LOG_FILE):
        df.to_excel(LOG_FILE, index=False)
    else:
        old_df = pd.read_excel(LOG_FILE)
        pd.concat([old_df, df], ignore_index=True).to_excel(LOG_FILE, index=False)

# View history

def view_history(email):
    if os.path.exists(LOG_FILE):
        df = pd.read_excel(LOG_FILE)
        if email == DEVELOPER_EMAIL:
            return df
        else:
            return df[df['Email'] == email]
    else:
        return pd.DataFrame()

# Generate PDF report
def generate_pdf(input_data, prediction, shap_percents, suggestion_text):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    pdf.cell(200, 10, txt="Concrete Strength Prediction Report", ln=True, align="C")
    pdf.ln(10)

    for key, value in input_data.items():
        pdf.cell(200, 10, txt=f"{key}: {value}", ln=True)

    pdf.ln(5)
    pdf.cell(200, 10, txt=f"Predicted Strength: {prediction:.2f} MPa", ln=True)
    pdf.ln(10)

    pdf.cell(200, 10, txt="Constituent Contribution (%):", ln=True)
    for name, val in shap_percents.items():
        pdf.cell(200, 10, txt=f"{name}: {val:.2f}%", ln=True)

    pdf.ln(10)
    pdf.multi_cell(200, 10, txt="Suggestions:\n" + suggestion_text)

    file_name = f"Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf.output(file_name)
    return file_name

# Suggestion generator
def get_suggestions(shap_values):
    top_features = shap_values.abs.mean(0).argsort()[-3:][::-1]
    suggestions = ""
    for i in top_features:
        feature = features[i]
        suggestions += f"- {feature} significantly affects strength. Adjust its proportion carefully.\n"
    return suggestions

# Prediction function
def predict_strength(email, cement, sand, coarse_agg, water, superplasticizer, fly_ash, slag, age):
    inputs = {
        "Cement": cement,
        "Sand": sand,
        "Coarse Aggregate": coarse_agg,
        "Water": water,
        "Superplasticizer": superplasticizer,
        "Fly Ash": fly_ash,
        "Slag": slag,
        "Age": age
    }
    df = pd.DataFrame([inputs])
    scaled = scaler.transform(df)

    pred1 = model.predict(scaled)[0]
    rf.fit(df, [pred1])
    gb.fit(df, [pred1])
    lr.fit(df, [pred1])
    pred2 = rf.predict(df)[0]
    pred3 = gb.predict(df)[0]
    pred4 = lr.predict(df)[0]
    final_pred = np.mean([pred1, pred2, pred3, pred4])

    log_prediction(email, inputs, final_pred)

    # SHAP Explanation
    shap_values = get_shap_values(df)
    shap_vals = shap_values.values[0]
    shap_sum = np.sum(np.abs(shap_vals))
    shap_percents = {features[i]: 100 * abs(shap_vals[i]) / shap_sum for i in range(len(features))}

    fig, ax = plt.subplots()
    plt.pie(shap_percents.values(), labels=shap_percents.keys(), autopct='%1.1f%%')
    plt.title('Strength Contribution by Constituents')
    plt.savefig("shap_pie.png")

    suggestion_text = get_suggestions(shap_values)
    report_path = generate_pdf(inputs, final_pred, shap_percents, suggestion_text)

    return final_pred, shap_percents, "shap_pie.png", report_path

# Bulk upload handler
def bulk_predict(email, file):
    df = pd.read_excel(file.name)
    results = []
    for i, row in df.iterrows():
        pred, *_ = predict_strength(email, *row.values)
        row_result = row.to_dict()
        row_result["Prediction"] = pred
        results.append(row_result)
    return pd.DataFrame(results)

# Gradio Interface
with gr.Blocks() as app:
    gr.Markdown("# AI Concrete Strength Predictor")

    email = gr.Textbox(label="Your Email")

    with gr.Tab("Manual Prediction"):
        inputs = [
            gr.Number(label=label) for label in features
        ]
        predict_btn = gr.Button("Predict")
        result = gr.Textbox(label="Predicted Strength (MPa)")
        shap_txt = gr.JSON(label="Contribution %")
        shap_img = gr.Image(label="Contribution Pie Chart")
        pdf_out = gr.File(label="Download Report")

    with gr.Tab("Bulk Prediction"):
        file_input = gr.File(label="Upload Excel (.xlsx)")
        bulk_out = gr.Dataframe()
        bulk_btn = gr.Button("Predict Bulk")

    with gr.Tab("View History"):
        view_btn = gr.Button("View My History")
        history_table = gr.Dataframe()

    predict_btn.click(fn=predict_strength, inputs=[email] + inputs, outputs=[result, shap_txt, shap_img, pdf_out])
    bulk_btn.click(fn=bulk_predict, inputs=[email, file_input], outputs=bulk_out)
    view_btn.click(fn=view_history, inputs=[email], outputs=history_table)

app.launch()
